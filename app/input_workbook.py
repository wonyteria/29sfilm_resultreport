from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GUIDE_FILL = PatternFill("solid", fgColor="D9EAF7")


SHEETS = {
    "기본정보": ["항목", "값", "상태", "근거파일", "비고"],
    "전체일정": ["순서", "구분", "시작일", "종료일", "내용", "보고서표시", "상태", "근거파일"],
    "홍보실적": ["채널", "세부유형", "시작일", "종료일", "횟수", "대상수", "도달", "노출", "조회", "클릭", "참여", "기준일", "상태", "근거파일", "URL"],
    "출품현황": ["기준일", "구분", "출품수", "누적여부", "상태", "근거파일"],
    "심사정보": ["단계", "시작일", "종료일", "방식", "장소", "심사위원", "소속", "직책", "상태", "근거파일"],
    "수상작": ["순서", "상격", "부문", "수상자", "작품명", "상금", "훈격", "URL", "시놉시스", "대표이미지", "상태"],
    "시상식": ["항목", "값", "상태", "근거파일", "비고"],
    "참석자후기": ["유형", "이름·출처", "내용", "URL", "이미지", "채택여부", "상태"],
    "총평": ["항목", "내용", "상태", "비고"],
    "페이지구성": ["대분류", "모듈키", "페이지명", "구성결정", "판단근거", "담당자메모"],
}


BASIC_ROWS = [
    "행사유형", "사업명", "회차", "주최", "주관", "후원", "사업시작일", "사업종료일", "공모시작일", "공모종료일",
    "사업목적", "공모주제", "총상금", "시상팀수", "보고서기준일", "작성자",
]

CEREMONY_ROWS = ["시상식명", "일시", "장소", "진행방식", "사회자", "참석인원", "주요내빈", "온라인중계URL"]
SUMMARY_ROWS = ["주요성과", "정량성과", "주최사목표달성", "참가자반응", "운영상특징", "개선점", "차기사업제안", "필수표현", "금지표현", "AI초안", "최종승인본"]


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    widths = {1: 18, 2: 26, 3: 18, 4: 28, 5: 30}
    for index in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(index, 16)


def create_input_workbook(path: str | Path, event_type: str = "", report_plan=None) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["29초영화제·29역숏폼왕 결과보고 입력 파일"])
    guide.append(["파란색 제목 행과 시트 이름은 변경하지 마세요."])
    guide.append(["상태는 미입력, 자동추출, 확인필요, 담당자확인, 해당없음 중 하나를 사용하세요."])
    guide.append(["숫자·날짜·인명은 담당자확인 상태가 되어야 완성본에 반영됩니다."])
    guide.column_dimensions["A"].width = 100
    guide["A1"].font = Font(size=16, bold=True)
    guide["A1"].fill = GUIDE_FILL

    for title, headers in SHEETS.items():
        ws = wb.create_sheet(title)
        ws.append(headers)
        if title == "기본정보":
            for item in BASIC_ROWS:
                ws.append([item, event_type if item == "행사유형" else "", "미입력", "", ""])
        elif title == "시상식":
            for item in CEREMONY_ROWS:
                ws.append([item, "", "미입력", "", ""])
        elif title == "총평":
            for item in SUMMARY_ROWS:
                ws.append([item, "", "미입력", ""])
        elif title == "페이지구성" and report_plan:
            for item in report_plan:
                ws.append([item.section, item.key, item.label, item.decision, item.reason, ""])
        _format_sheet(ws)
    wb.save(path)
    return path
