import tempfile
import unittest
import zipfile
from pathlib import Path

from app.detection import detect_event_type
from app.folder_template import FOLDERS, create_folder_zip
from app.readiness import assess_readiness
from app.scanner import scan_project
from app.page_catalog import classify_slide, normalize
from app.report_plan import build_report_plan
from app.input_workbook import create_input_workbook
from openpyxl import load_workbook
from app.content_extract import extract_project_text
from app.project_state import save_project
from app.quality_gate import inspect_plan
from app.template_registry import recommend_templates


class CoreTests(unittest.TestCase):
    def test_extracts_csv_evidence(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "수상작.csv").write_text("부문,작품명\n일반부,테스트 작품", encoding="utf-8")
            evidence = extract_project_text(str(root), ["수상작.csv"])
            self.assertIn("테스트 작품", evidence[0].text)

    def test_project_state_and_quality_gate(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "인스타그램").mkdir()
            (root / "인스타그램" / "게시물.png").write_bytes(b"image")
            scan = scan_project(str(root))
            plan = build_report_plan(scan, "29초영화제")
            target = save_project(root / "project.json", str(root), "29초영화제", plan)
            self.assertTrue(target.exists())
            self.assertGreater(len(inspect_plan(plan)), 0)

    def test_template_recommendation_respects_event_type(self):
        catalog = Path(__file__).parents[1] / "page_catalog.json"
        choices = recommend_templates(catalog, "29역숏폼왕", {"cover", "sns", "award_detail"})
        self.assertEqual(choices[0].event_type, "29역숏폼왕")

    def test_zip_contains_standard_folders(self):
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "template.zip"
            create_folder_zip(target)
            with zipfile.ZipFile(target) as zf:
                names = set(zf.namelist())
            self.assertIn("01.최종기획안/", names)
            self.assertIn("99.최종/", names)
            self.assertIn("08.수상작/수상작_입력.csv", names)

    def test_scan_excludes_result_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "03.언론홍보").mkdir()
            (root / "03.언론홍보" / "기사.jpg").write_bytes(b"data")
            (root / "제12회 신한 29초영화제_결과보고서.pptx").write_bytes(b"draft")
            result = scan_project(str(root))
            self.assertEqual(result.total_files, 1)
            self.assertEqual(result.excluded_files, 1)

    def test_detects_event_type(self):
        detection = detect_event_type(r"Z:\2026\29역숏폼왕\정식품\06.결과보고")
        self.assertEqual(detection.event_type, "29역숏폼왕")
        self.assertEqual(detection.confidence, "높음")

    def test_page_module_classification(self):
        module = classify_slide(normalize("3-2. 심사 심사위원 심사 방식"), 10, 20)
        self.assertIsNotNone(module)
        self.assertEqual(module.key, "judging")

    def test_dynamic_plan_uses_evidence(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "인스타그램").mkdir()
            (root / "인스타그램" / "게시물.png").write_bytes(b"image")
            scan = scan_project(str(root))
            plan = {item.key: item for item in build_report_plan(scan, "29초영화제")}
            self.assertEqual(plan["sns"].decision, "반복 생성")
            self.assertEqual(plan["overview"].decision, "자동 포함")

    def test_input_workbook_has_required_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "input.xlsx"
            create_input_workbook(target, "29초영화제")
            wb = load_workbook(target, read_only=True)
            self.assertIn("기본정보", wb.sheetnames)
            self.assertIn("수상작", wb.sheetnames)
            self.assertIn("페이지구성", wb.sheetnames)
            wb.close()


if __name__ == "__main__":
    unittest.main()
